import openpyxl

path = r"C:\Users\ASUS\PycharmProjects\RobotTraining\login_details.xlsx"
workbook = openpyxl.load_workbook(path)

def fetchRows(sheet):
    sheet1 = workbook[sheet]
    number_of_rows = sheet1.max_row
    return number_of_rows

def fetchIdPwd(sheet, row, col):
    sheet1 = workbook[sheet]
    cell = sheet1.cell(row=int(row), column=int(col))
    return str(cell.value)
