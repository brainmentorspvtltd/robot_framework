import openpyxl

path = r"C:\Users\ASUS\PowerBI Datasets\EmployeesData.xlsx"
workbookObject = openpyxl.load_workbook(path)
sheet1 = workbookObject.active
# print(workbookObject.active)

# returns the sheet object
# sheet1 = workbookObject['Sheet1']
# print(sheet1)
print("Number of rows in sheet_1 :",sheet1.max_row)
print("Number of columns in sheet_1 :",sheet1.max_column)

# return cell object that you specify
# data = sheet1.cell(row=15, column=15)
# print("Row 1, Col 1 :", data.value)

sheet1['F1'] = 'Address'

data = ['delhi','pune','chennai','noida','lucknow']
for i, item in enumerate(data, start=2):
    sheet1['F{}'.format(i)] = item

workbookObject.save(path)
